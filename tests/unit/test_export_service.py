"""Tests for `app.core.services.export_service` — PR3 scope.

`ExportService`'s full behavioral contract is verified with a fake
`DocumentConverterProvider` (see spec's "Provider-Agnostic Testability"
requirement) — no real Office/COM is exercised here. `ComWordProvider`'s
own real-COM behavior is covered separately by
`tests/integration/test_com_word_provider.py`.
"""

from __future__ import annotations

import logging
from collections.abc import Callable
from pathlib import Path

import openpyxl
import pdfplumber
import pymupdf
import pytest
from docx import Document

from app.core.exceptions import (
    ConversionFallidaError,
    ConversorNoDisponibleError,
    EntradaInvalidaError,
    PDFCorruptoError,
    PDFSinTablasError,
    PDFSinTextoError,
)
from app.core.providers.azure_doc_converter_provider import AzureDocConverterProvider
from app.core.providers.com_word_provider import ComWordProvider
from app.core.providers.document_converter_provider import DocumentConverterProvider
from app.core.services.export_service import ExportService

_LOGGER_NAME = "app.core.services.export_service"


class FakeDocumentConverterProvider:
    """Configurable fake `DocumentConverterProvider` for `ExportService` tests.

    `behavior` selects what `convertir()` does:
        "succeed" -> writes a stub file at `output`, returns `output`
        "unavailable" -> raises `ConversorNoDisponibleError`
        "fail" -> raises a generic `RuntimeError` (simulates a raw
            COM/pywin32 failure `ExportService` must translate)
        "timeout" -> raises `TimeoutError`
        "not_implemented" -> raises `NotImplementedError` (Azure stub shape)
    """

    def __init__(self, behavior: str = "succeed") -> None:
        self.behavior = behavior
        self.called = False

    def convertir(self, source: Path, output: Path) -> Path:
        self.called = True
        if self.behavior == "succeed":
            output.write_bytes(b"%PDF-fake-output")
            return output
        if self.behavior == "unavailable":
            raise ConversorNoDisponibleError("fake provider unavailable")
        if self.behavior == "fail":
            raise RuntimeError("fake raw conversion failure")
        if self.behavior == "timeout":
            raise TimeoutError("fake conversion timed out")
        if self.behavior == "not_implemented":
            raise NotImplementedError("fake provider not implemented")
        raise AssertionError(f"unknown behavior: {self.behavior!r}")

    def esta_disponible(self) -> tuple[bool, str]:
        return True, "fake provider always available"


class TestProviderSelection:
    def test_default_provider_is_com_when_env_unset(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.delenv("DOC_CONVERTER_PROVIDER", raising=False)

        service = ExportService()

        assert isinstance(service._provider, ComWordProvider)

    def test_env_selects_azure_provider(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.setenv("DOC_CONVERTER_PROVIDER", "azure")

        service = ExportService()

        assert isinstance(service._provider, AzureDocConverterProvider)

    def test_env_change_after_construction_is_ignored(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.setenv("DOC_CONVERTER_PROVIDER", "com")
        service = ExportService()

        monkeypatch.setenv("DOC_CONVERTER_PROVIDER", "azure")

        assert isinstance(service._provider, ComWordProvider)

    def test_explicit_provider_overrides_factory(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.setenv("DOC_CONVERTER_PROVIDER", "azure")
        fake = FakeDocumentConverterProvider("succeed")

        service = ExportService(provider=fake)

        assert service._provider is fake


class TestValidation:
    def test_rejects_non_docx_extension_before_provider_invoked(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        fake = FakeDocumentConverterProvider("succeed")
        service = ExportService(provider=fake)
        source = tmp_path / "not_a_docx.txt"
        source.write_text("hello")

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.convertir(source, tmp_path / "out.pdf")

        assert fake.called is False
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert len(warning_records) == 1
        assert "not_a_docx.txt" in warning_records[0].getMessage()
        assert str(tmp_path) not in warning_records[0].getMessage()

    @pytest.mark.parametrize("suffix", [".doc", ".docm"])
    def test_rejects_other_word_formats_out_of_v1_scope(
        self, tmp_path: Path, suffix: str
    ) -> None:
        fake = FakeDocumentConverterProvider("succeed")
        service = ExportService(provider=fake)
        source = tmp_path / f"legacy{suffix}"
        source.write_text("not actually a docx, extension is what's tested")

        with pytest.raises(EntradaInvalidaError):
            service.convertir(source, tmp_path / "out.pdf")

        assert fake.called is False

    def test_accepts_uppercase_docx_extension(
        self, tmp_path: Path, valid_docx_factory: Callable[..., Path]
    ) -> None:
        fake = FakeDocumentConverterProvider("succeed")
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")
        uppercased = source.with_suffix(".DOCX")
        source.rename(uppercased)

        result = service.convertir(uppercased, tmp_path / "out.pdf")

        assert fake.called is True
        assert result == tmp_path / "out.pdf"

    def test_rejects_missing_source(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        fake = FakeDocumentConverterProvider("succeed")
        service = ExportService(provider=fake)
        source = tmp_path / "missing.docx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.convertir(source, tmp_path / "out.pdf")

        assert fake.called is False

    def test_rejects_empty_source(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        fake = FakeDocumentConverterProvider("succeed")
        service = ExportService(provider=fake)
        source = tmp_path / "empty.docx"
        source.write_bytes(b"")

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.convertir(source, tmp_path / "out.pdf")

        assert fake.called is False

    def test_rejects_uncreatable_output_dir(
        self,
        tmp_path: Path,
        valid_docx_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        fake = FakeDocumentConverterProvider("succeed")
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")
        blocked_by_file = tmp_path / "blocked_by_file"
        blocked_by_file.write_bytes(b"not a directory")
        output = blocked_by_file / "sub" / "out.pdf"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.convertir(source, output)

        assert fake.called is False


class TestHappyPath:
    def test_convertir_returns_output_path_and_logs(
        self,
        tmp_path: Path,
        valid_docx_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        fake = FakeDocumentConverterProvider("succeed")
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")
        output = tmp_path / "out.pdf"

        with caplog.at_level(logging.INFO, logger=_LOGGER_NAME):
            result = service.convertir(source, output)

        assert result == output
        assert output.exists()
        assert fake.called is True
        info_records = [r for r in caplog.records if r.levelno == logging.INFO]
        assert any("start" in r.getMessage() for r in info_records)
        assert any("ok" in r.getMessage() for r in info_records)
        for record in info_records:
            assert str(tmp_path) not in record.getMessage()


class TestProviderErrorTranslation:
    def test_unavailable_provider_raises_conversor_no_disponible(
        self,
        tmp_path: Path,
        valid_docx_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        fake = FakeDocumentConverterProvider("unavailable")
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")

        with pytest.raises(ConversorNoDisponibleError):
            service.convertir(source, tmp_path / "out.pdf")

    def test_generic_provider_failure_raises_conversion_fallida(
        self, tmp_path: Path, valid_docx_factory: Callable[..., Path]
    ) -> None:
        fake = FakeDocumentConverterProvider("fail")
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")

        with pytest.raises(ConversionFallidaError):
            service.convertir(source, tmp_path / "out.pdf")

    def test_timeout_raises_conversion_fallida(
        self, tmp_path: Path, valid_docx_factory: Callable[..., Path]
    ) -> None:
        fake = FakeDocumentConverterProvider("timeout")
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")

        with pytest.raises(ConversionFallidaError):
            service.convertir(source, tmp_path / "out.pdf")

    def test_provider_failure_logs_warning_filename_only(
        self,
        tmp_path: Path,
        valid_docx_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        fake = FakeDocumentConverterProvider("fail")
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            ConversionFallidaError
        ):
            service.convertir(source, tmp_path / "out.pdf")

        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert any("doc.docx" in r.getMessage() for r in warning_records)
        for record in warning_records:
            assert str(tmp_path) not in record.getMessage()

    def test_raw_provider_exception_never_escapes(
        self, tmp_path: Path, valid_docx_factory: Callable[..., Path]
    ) -> None:
        fake = FakeDocumentConverterProvider("fail")
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")

        with pytest.raises(ConversionFallidaError) as exc_info:
            service.convertir(source, tmp_path / "out.pdf")

        assert type(exc_info.value) is ConversionFallidaError

    @pytest.mark.parametrize("behavior", ["fail", "timeout"])
    def test_provider_failure_leaves_no_partial_output_file(
        self, tmp_path: Path, valid_docx_factory: Callable[..., Path], behavior: str
    ) -> None:
        """No partial/corrupt PDF survives a provider failure.

        `_make_output_dir` runs before the provider call (design's
        documented, PDFService-consistent tradeoff — see
        `ExportService._make_output_dir`'s docstring), so the output
        *directory* may exist afterward; what must NOT happen is a
        partial or stale file at `output` itself.
        """
        fake = FakeDocumentConverterProvider(behavior)
        service = ExportService(provider=fake)
        source = valid_docx_factory("doc.docx")
        output = tmp_path / "out.pdf"

        with pytest.raises(ConversionFallidaError):
            service.convertir(source, output)

        assert not output.exists()


class TestAzureStub:
    def test_azure_provider_convertir_propagates_not_implemented(
        self, tmp_path: Path, valid_docx_factory: Callable[..., Path]
    ) -> None:
        service = ExportService(provider=AzureDocConverterProvider())
        source = valid_docx_factory("doc.docx")

        with pytest.raises(NotImplementedError):
            service.convertir(source, tmp_path / "out.pdf")

    def test_azure_selected_via_env_propagates_not_implemented(
        self,
        tmp_path: Path,
        valid_docx_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        monkeypatch.setenv("DOC_CONVERTER_PROVIDER", "azure")
        service = ExportService()
        source = valid_docx_factory("doc.docx")

        with pytest.raises(NotImplementedError):
            service.convertir(source, tmp_path / "out.pdf")


def test_fake_provider_satisfies_the_protocol() -> None:
    fake = FakeDocumentConverterProvider("succeed")

    assert isinstance(fake, DocumentConverterProvider)


class TestPdfAWordHappyPath:
    """`ExportService.pdf_a_word` — `.pdf` -> `.docx`, per `sdd/pdf-to-word`."""

    def test_returns_output_path_and_content_matches_source_text(
        self,
        tmp_path: Path,
        native_text_pdf_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        text = "Documento con texto nativo de prueba."
        source = native_text_pdf_factory("native.pdf", text)
        output = tmp_path / "out.docx"

        with caplog.at_level(logging.INFO, logger=_LOGGER_NAME):
            result = service.pdf_a_word(source, output)

        assert result == output
        assert output.exists()
        document = Document(str(output))
        full_text = "\n".join(paragraph.text for paragraph in document.paragraphs)
        assert text in full_text
        info_records = [r for r in caplog.records if r.levelno == logging.INFO]
        assert any("start" in r.getMessage() for r in info_records)
        assert any("ok" in r.getMessage() for r in info_records)
        for record in info_records:
            assert str(tmp_path) not in record.getMessage()


class TestPdfAWordValidation:
    def test_rejects_non_pdf_extension_before_processing(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = tmp_path / "not_a_pdf.txt"
        source.write_text("hello")
        output = tmp_path / "out.docx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.pdf_a_word(source, output)

        assert not output.exists()
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert len(warning_records) == 1
        assert "not_a_pdf.txt" in warning_records[0].getMessage()
        assert str(tmp_path) not in warning_records[0].getMessage()

    def test_rejects_missing_source(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = tmp_path / "missing.pdf"
        output = tmp_path / "out.docx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.pdf_a_word(source, output)

        assert not output.exists()

    def test_rejects_empty_source(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = tmp_path / "empty.pdf"
        source.write_bytes(b"")
        output = tmp_path / "out.docx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.pdf_a_word(source, output)

        assert not output.exists()

    def test_rejects_uncreatable_output_dir(
        self,
        tmp_path: Path,
        native_text_pdf_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = native_text_pdf_factory("native.pdf", "Some real text content here.")
        blocked_by_file = tmp_path / "blocked_by_file"
        blocked_by_file.write_bytes(b"not a directory")
        output = blocked_by_file / "sub" / "out.docx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.pdf_a_word(source, output)


class TestPdfAWordScannedRejection:
    """The CONFIRMED-critical path — `pdf2docx` silently "succeeds" on
    scanned input, so the upfront `pymupdf`-based detection gate MUST run
    and MUST prevent `Converter` from ever being constructed."""

    def test_rejects_image_only_pdf_before_converter_is_ever_invoked(
        self,
        tmp_path: Path,
        image_only_pdf_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        calls: list[str] = []

        class _SpyConverter:
            def __init__(self, *args: object, **kwargs: object) -> None:
                calls.append("constructed")

            def convert(self, *args: object, **kwargs: object) -> None:
                calls.append("convert")

            def close(self) -> None:
                calls.append("close")

        monkeypatch.setattr("app.core.services.export_service.Converter", _SpyConverter)
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = image_only_pdf_factory("scanned.pdf")
        output = tmp_path / "out.docx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            PDFSinTextoError
        ) as exc_info:
            service.pdf_a_word(source, output)

        assert type(exc_info.value) is PDFSinTextoError
        assert "OCR" in str(exc_info.value)
        assert calls == []
        assert not output.exists()
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert len(warning_records) == 1
        assert "scanned.pdf" in warning_records[0].getMessage()
        assert str(tmp_path) not in warning_records[0].getMessage()

    def test_text_only_on_a_later_page_still_counts_as_native_text(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Regression: the detection gate sums non-whitespace chars across
        EVERY page (`sum(... for page in doc)`), not just the first. A
        regression to checking only page 1 would wrongly reject a PDF
        whose real text lives on a later page."""
        calls: list[str] = []

        class _SpyConverter:
            def __init__(self, *args: object, **kwargs: object) -> None:
                calls.append("constructed")

            def convert(self, *args: object, **kwargs: object) -> None:
                calls.append("convert")
                Path(args[0]).write_bytes(b"fake docx bytes")

            def close(self) -> None:
                calls.append("close")

        monkeypatch.setattr("app.core.services.export_service.Converter", _SpyConverter)

        source = tmp_path / "mixed_pages.pdf"
        doc = pymupdf.open()
        doc.new_page(width=612, height=792)  # page 1: blank, no text
        page2 = doc.new_page(width=612, height=792)
        page2.insert_text((72, 72), "Real extractable text lives only on page two.")
        doc.save(source)
        doc.close()

        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        output = tmp_path / "out.docx"

        result = service.pdf_a_word(source, output)

        assert result == output
        assert "constructed" in calls
        assert "convert" in calls


class TestPdfAWordCorruptPdf:
    def test_rejects_corrupt_pdf_at_open_time(
        self,
        tmp_path: Path,
        corrupt_pdf_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = corrupt_pdf_factory("corrupt.pdf")
        output = tmp_path / "out.docx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            PDFCorruptoError
        ) as exc_info:
            service.pdf_a_word(source, output)

        assert type(exc_info.value) is PDFCorruptoError
        assert not output.exists()
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert any("corrupt.pdf" in r.getMessage() for r in warning_records)
        for record in warning_records:
            assert str(tmp_path) not in record.getMessage()

    def test_rejects_corrupt_pdf_at_converter_construction_time(
        self,
        tmp_path: Path,
        native_text_pdf_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Confirmed distinct failure point: `Converter(str(source))` (the
        CONSTRUCTOR) can itself raise `pymupdf.FileDataError`, separately
        from the earlier `pymupdf.open()` detection step — both must be
        caught by the same `_translate_pdf_open_errors` boundary."""

        class _RaisingConverter:
            def __init__(self, *args: object, **kwargs: object) -> None:
                raise pymupdf.FileDataError("simulated Converter() construction failure")

        monkeypatch.setattr("app.core.services.export_service.Converter", _RaisingConverter)
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = native_text_pdf_factory("native.pdf", "Some real text content here.")
        output = tmp_path / "out.docx"

        with pytest.raises(PDFCorruptoError) as exc_info:
            service.pdf_a_word(source, output)

        assert type(exc_info.value) is PDFCorruptoError
        assert not output.exists()


class TestPdfAWordConversionFailure:
    def test_leaves_no_partial_output_on_conversion_failure(
        self,
        tmp_path: Path,
        native_text_pdf_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """No natural way to make `pdf2docx` fail on an otherwise-valid
        native-text PDF, so `Converter.convert` is monkeypatched to raise
        after writing a partial file — matching this codebase's existing
        precedent for hard-to-trigger real failures (e.g.
        `ComWordProvider`'s timeout tests)."""

        close_calls: list[bool] = []

        class _FailingConverter:
            def __init__(self, *args: object, **kwargs: object) -> None:
                pass

            def convert(self, docx_filename: str, *args: object, **kwargs: object) -> None:
                Path(docx_filename).write_bytes(b"partial, mid-conversion garbage")
                raise RuntimeError("simulated pdf2docx conversion failure")

            def close(self) -> None:
                close_calls.append(True)

        monkeypatch.setattr("app.core.services.export_service.Converter", _FailingConverter)
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = native_text_pdf_factory("native.pdf", "Some real text content here.")
        output = tmp_path / "out.docx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            ConversionFallidaError
        ) as exc_info:
            service.pdf_a_word(source, output)

        assert type(exc_info.value) is ConversionFallidaError
        assert not output.exists()
        # Regression: cv.close() must run even when .convert() raises, or
        # pdf2docx's underlying pymupdf handle on `source` leaks.
        assert close_calls == [True]
        # Regression: nothing outside the temp working area should survive
        # a failed conversion — no stray "*.docx" siblings of `output`.
        assert list(tmp_path.glob("*.docx")) == []
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert any("native.pdf" in r.getMessage() for r in warning_records)
        for record in warning_records:
            assert str(tmp_path) not in record.getMessage()

    def test_preexisting_output_file_survives_a_conversion_failure_untouched(
        self,
        tmp_path: Path,
        native_text_pdf_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Regression: an earlier revision unconditionally unlinked
        `output` on any conversion failure. If `output` already pointed
        at a pre-existing file (e.g. the caller overwriting a prior
        successful run), that would DELETE the caller's original file
        even though nothing succeeded — worse than leaving a partial
        write. Converting into a temp path first (per the fix) means
        `output` must never be touched at all on failure."""

        class _FailingConverter:
            def __init__(self, *args: object, **kwargs: object) -> None:
                pass

            def convert(self, docx_filename: str, *args: object, **kwargs: object) -> None:
                raise RuntimeError("simulated pdf2docx conversion failure")

            def close(self) -> None:
                pass

        monkeypatch.setattr("app.core.services.export_service.Converter", _FailingConverter)
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = native_text_pdf_factory("native.pdf", "Some real text content here.")
        output = tmp_path / "out.docx"
        original_bytes = b"a pre-existing docx the caller does not want deleted"
        output.write_bytes(original_bytes)

        with pytest.raises(ConversionFallidaError):
            service.pdf_a_word(source, output)

        assert output.exists()
        assert output.read_bytes() == original_bytes


class TestPdfAExcelHappyPath:
    """`ExportService.pdf_a_excel` — `.pdf` -> `.xlsx`, per `sdd/pdf-to-excel`."""

    def test_multi_table_produces_multiple_worksheets(
        self,
        tmp_path: Path,
        table_pdf_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        tables = [
            [["Header1", "Header2"], ["A1", "B1"], ["A2", "B2"]],
            [["X", "Y", "Z"], ["1", "2", "3"]],
        ]
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = table_pdf_factory("tables.pdf", tables)
        output = tmp_path / "out.xlsx"

        with caplog.at_level(logging.INFO, logger=_LOGGER_NAME):
            result = service.pdf_a_excel(source, output)

        assert result == output
        assert output.exists()
        workbook = openpyxl.load_workbook(output)
        assert workbook.sheetnames == ["Tabla1", "Tabla2"]
        sheet1_rows = [list(row) for row in workbook["Tabla1"].iter_rows(values_only=True)]
        sheet2_rows = [list(row) for row in workbook["Tabla2"].iter_rows(values_only=True)]
        assert sheet1_rows == tables[0]
        assert sheet2_rows == tables[1]
        info_records = [r for r in caplog.records if r.levelno == logging.INFO]
        assert any("start" in r.getMessage() for r in info_records)
        assert any("ok" in r.getMessage() for r in info_records)
        for record in info_records:
            assert str(tmp_path) not in record.getMessage()

    def test_three_tables_produce_three_correctly_named_worksheets(
        self, tmp_path: Path, table_pdf_factory: Callable[..., Path]
    ) -> None:
        """Regression for the `index == 1` / `else` sheet-naming branch:
        the 2-table happy-path test above can't distinguish "renamed
        .active for table 1, create_sheet for table 2" from a subtly
        wrong off-by-one that would only surface at table 3+."""
        tables = [
            [["Alpha1", "Alpha2"]],
            [["Beta1", "Beta2"]],
            [["Gamma1", "Gamma2"]],
        ]
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = table_pdf_factory("three_tables.pdf", tables)
        output = tmp_path / "out.xlsx"

        service.pdf_a_excel(source, output)

        workbook = openpyxl.load_workbook(output)
        assert workbook.sheetnames == ["Tabla1", "Tabla2", "Tabla3"]
        for i, expected in enumerate(tables, start=1):
            rows = [list(row) for row in workbook[f"Tabla{i}"].iter_rows(values_only=True)]
            assert rows == expected


class TestPdfAExcelValidation:
    def test_rejects_non_pdf_extension_before_processing(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = tmp_path / "not_a_pdf.txt"
        source.write_text("hello")
        output = tmp_path / "out.xlsx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            EntradaInvalidaError
        ):
            service.pdf_a_excel(source, output)

        assert not output.exists()
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert len(warning_records) == 1
        assert "not_a_pdf.txt" in warning_records[0].getMessage()
        assert str(tmp_path) not in warning_records[0].getMessage()

    def test_rejects_missing_source(self, tmp_path: Path) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = tmp_path / "missing.pdf"
        output = tmp_path / "out.xlsx"

        with pytest.raises(EntradaInvalidaError):
            service.pdf_a_excel(source, output)

        assert not output.exists()

    def test_rejects_empty_source(self, tmp_path: Path) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = tmp_path / "empty.pdf"
        source.write_bytes(b"")
        output = tmp_path / "out.xlsx"

        with pytest.raises(EntradaInvalidaError):
            service.pdf_a_excel(source, output)

        assert not output.exists()

    def test_rejects_uncreatable_output_dir(
        self, tmp_path: Path, table_pdf_factory: Callable[..., Path]
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = table_pdf_factory("tables.pdf")
        blocked_by_file = tmp_path / "blocked_by_file"
        blocked_by_file.write_bytes(b"not a directory")
        output = blocked_by_file / "sub" / "out.xlsx"

        with pytest.raises(EntradaInvalidaError):
            service.pdf_a_excel(source, output)

        assert not output.exists()


class TestPdfAExcelZeroTableRejection:
    """`PDFSinTablasError` — distinct from `PDFSinTextoError`, no OCR suggestion."""

    def test_rejects_prose_pdf_before_workbook_is_ever_constructed(
        self,
        tmp_path: Path,
        native_text_pdf_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        calls: list[str] = []

        class _SpyWorkbook:
            def __init__(self, *args: object, **kwargs: object) -> None:
                calls.append("constructed")

        monkeypatch.setattr(openpyxl, "Workbook", _SpyWorkbook)
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = native_text_pdf_factory(
            "prose.pdf", "Just a plain paragraph of prose, no grid or table at all."
        )
        output = tmp_path / "out.xlsx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            PDFSinTablasError
        ) as exc_info:
            service.pdf_a_excel(source, output)

        assert type(exc_info.value) is PDFSinTablasError
        assert "OCR" not in str(exc_info.value)
        assert calls == []
        assert not output.exists()
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert len(warning_records) == 1
        assert "prose.pdf" in warning_records[0].getMessage()
        assert str(tmp_path) not in warning_records[0].getMessage()


class TestPdfAExcelScannedRejection:
    """A scanned/image-only PDF must raise `PDFSinTextoError`, NOT
    `PDFSinTablasError` — proves the text-gate routes before `pdfplumber`
    ever runs (mirrors `pdf_a_word`'s scanned-rejection precedent)."""

    def test_rejects_image_only_pdf_before_pdfplumber_is_ever_invoked(
        self,
        tmp_path: Path,
        image_only_pdf_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        calls: list[str] = []

        def _raise_if_called(*args: object, **kwargs: object) -> None:
            calls.append("opened")
            raise AssertionError("pdfplumber.open must not be called on a scanned PDF")

        monkeypatch.setattr(pdfplumber, "open", _raise_if_called)
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = image_only_pdf_factory("scanned.pdf")
        output = tmp_path / "out.xlsx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            PDFSinTextoError
        ) as exc_info:
            service.pdf_a_excel(source, output)

        assert type(exc_info.value) is PDFSinTextoError
        assert "OCR" in str(exc_info.value)
        assert calls == []
        assert not output.exists()
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert len(warning_records) == 1
        assert "scanned.pdf" in warning_records[0].getMessage()
        assert str(tmp_path) not in warning_records[0].getMessage()


class TestPdfAExcelCorruptPdf:
    def test_rejects_corrupt_pdf_at_open_time(
        self,
        tmp_path: Path,
        corrupt_pdf_factory: Callable[..., Path],
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = corrupt_pdf_factory("corrupt.pdf")
        output = tmp_path / "out.xlsx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            PDFCorruptoError
        ) as exc_info:
            service.pdf_a_excel(source, output)

        assert type(exc_info.value) is PDFCorruptoError
        assert not output.exists()
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert any("corrupt.pdf" in r.getMessage() for r in warning_records)
        for record in warning_records:
            assert str(tmp_path) not in record.getMessage()


class TestPdfAExcelExtractionFailure:
    def test_openpyxl_save_failure_raises_conversion_fallida_no_output(
        self,
        tmp_path: Path,
        table_pdf_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """No natural way to make `openpyxl` fail on an otherwise-valid
        table PDF, so `Workbook.save` is monkeypatched to raise — matching
        `pdf_a_word`'s established precedent for hard-to-trigger real
        failures."""

        def _raise_save(self: object, *args: object, **kwargs: object) -> None:
            raise RuntimeError("simulated openpyxl save failure")

        monkeypatch.setattr(openpyxl.Workbook, "save", _raise_save)
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = table_pdf_factory("tables.pdf")
        output = tmp_path / "out.xlsx"

        with caplog.at_level(logging.WARNING, logger=_LOGGER_NAME), pytest.raises(
            ConversionFallidaError
        ) as exc_info:
            service.pdf_a_excel(source, output)

        assert type(exc_info.value) is ConversionFallidaError
        assert not output.exists()
        # Regression: the temp `.xlsx` must be cleaned up, not left orphaned.
        assert list(tmp_path.glob("*.xlsx")) == []
        warning_records = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert any("tables.pdf" in r.getMessage() for r in warning_records)
        for record in warning_records:
            assert str(tmp_path) not in record.getMessage()

    def test_preexisting_output_file_survives_an_extraction_failure_untouched(
        self,
        tmp_path: Path,
        table_pdf_factory: Callable[..., Path],
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Mandatory validate-then-write regression, mirroring
        `pdf_a_word`'s CRITICAL-bug-fix test: a pre-existing `output` must
        never be deleted or overwritten on any failure path, because the
        workbook is built into a temp file first and only moved to
        `output` via `os.replace()` on success."""

        def _raise_save(self: object, *args: object, **kwargs: object) -> None:
            raise RuntimeError("simulated openpyxl save failure")

        monkeypatch.setattr(openpyxl.Workbook, "save", _raise_save)
        service = ExportService(provider=FakeDocumentConverterProvider("succeed"))
        source = table_pdf_factory("tables.pdf")
        output = tmp_path / "out.xlsx"
        original_bytes = b"a pre-existing xlsx the caller does not want deleted"
        output.write_bytes(original_bytes)

        with pytest.raises(ConversionFallidaError):
            service.pdf_a_excel(source, output)

        assert output.exists()
        assert output.read_bytes() == original_bytes
